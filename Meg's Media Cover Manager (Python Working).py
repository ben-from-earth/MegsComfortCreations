import os
import io
import math
import json
import shutil
import requests
import openpyxl
import tkinter as tk
import tkinter.font as tkFont
from tkinter import messagebox, filedialog
from datetime import date
from PIL import Image, ImageTk

# ----------------------- GLOBAL CONSTANTS & GLOBAL VARIABLES -----------------------
SCALE = 1.5  # Global scale factor for UI scaling (1.5 means 50% larger)

API_KEY = "AIzaSyCtoXKuRUP5p0Xrk21635t67OA6MxFLay4"
CX = "e1e30c1aaa513492b"

# Folder paths
GATHERED_IMAGES_FOLDER = "Gathered Images"
DATABASE_FOLDER = "database"
UI_FOLDER = "UI Stuff"

# UI files
BACKGROUND_IMAGE_FILE = os.path.join(UI_FOLDER, "Flower Background.png")
DAILY_QUERIES_FILE = os.path.join(UI_FOLDER, "daily_queries.json")
BOOK_METADATA_FILE = os.path.join(UI_FOLDER, "book_metadata.json")

# Category mapping (used for folder names and database lookups)
CATEGORY_MAPPING = {
    "Books": "Books",
    "Movies": "Movies",
    "Video Games": "Video Games",
    "Music Records": "Albums"
}

# Global variable to store original (spaced) book inputs for later metadata prompting
ORIGINAL_BOOK_INPUTS = {}

# ----------------------- UTILITY FUNCTIONS -----------------------
def normalize_title(title):
    """
    Normalize a title by stripping extra whitespace and removing a leading 'The ' (case-insensitive).
    """
    title = title.strip()
    if title.lower().startswith("the "):
        title = title[4:]
    return title.strip()

def composite_key(title, author):
    """
    Returns the composite key used to index the metadata.
    Both title and author are normalized by removing all spaces.
    Example: "My Book" and "Author Name" become "MyBook_AuthorName".
    """
    normalized_title = "".join(normalize_title(title).split())
    normalized_author = "".join(author.strip().split())
    return f"{normalized_title}_{normalized_author}" if normalized_author else normalized_title

def find_duplicates(base_filename, dest_folder):
    """
    Searches for duplicate files in the destination folder based on the provided base filename.
    Returns a list of full paths for matching duplicates.
    """
    duplicates = []
    for file in os.listdir(dest_folder):
        name, ext = os.path.splitext(file)
        if name.lower().startswith(base_filename.lower()):
            duplicates.append(os.path.join(dest_folder, file))
    return duplicates

# ----------------------- PERSISTENCE FUNCTIONS -----------------------
def load_daily_queries():
    if not os.path.exists(DAILY_QUERIES_FILE):
        return 0
    try:
        with open(DAILY_QUERIES_FILE, "r") as f:
            data = json.load(f)
        return data.get("count", 0) if data.get("date") == str(date.today()) else 0
    except Exception as e:
        print("Error reading daily queries file:", e)
        return 0

def save_daily_queries(count):
    data = {"date": str(date.today()), "count": count}
    try:
        with open(DAILY_QUERIES_FILE, "w") as f:
            json.dump(data, f)
    except Exception as e:
        print("Error saving daily queries file:", e)

daily_queries = load_daily_queries()

def load_book_metadata():
    if not os.path.exists(BOOK_METADATA_FILE):
        return {}
    try:
        with open(BOOK_METADATA_FILE, "r") as f:
            return json.load(f)
    except Exception as e:
        print("Error loading book metadata:", e)
        return {}

def save_book_metadata(metadata):
    try:
        with open(BOOK_METADATA_FILE, "w") as f:
            json.dump(metadata, f, indent=4)
    except Exception as e:
        print("Error saving book metadata:", e)

def get_saved_genres():
    """
    Extracts and returns a sorted list of all genres found in the book metadata.
    """
    metadata = load_book_metadata()
    genres = set()
    for book in metadata.values():
        for g in book.get("genres", []):
            genres.add(g)
    return sorted(list(genres))

# ----------------------- GUI HELPER FUNCTIONS -----------------------
def load_background_image(file_path, width, height):
    """
    Loads and resizes an image to be used as a background.
    """
    try:
        img = Image.open(file_path)
        img = img.resize((width, height), Image.Resampling.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception as e:
        print(f"Could not load background image: {e}")
        return None

def apply_background(window, width=None, height=None):
    """
    Applies a background image to a given Tkinter window.
    If the image fails to load, sets a pink background.
    """
    window.update_idletasks()
    if width is None:
        width = window.winfo_width() or window.winfo_reqwidth()
    if height is None:
        height = window.winfo_height() or window.winfo_reqheight()
    bg_img = load_background_image(BACKGROUND_IMAGE_FILE, width, height)
    if bg_img:
        bg_label = tk.Label(window, image=bg_img)
        bg_label.place(x=0, y=0, relwidth=1, relheight=1)
        window.bg_img = bg_img  # Keep a reference to prevent garbage collection
    else:
        window.configure(bg="#f7acb0")
    return bg_img

def open_file_explorer(folder_path):
    """
    Opens the system file explorer at the given folder path.
    """
    if os.name == 'nt':
        os.startfile(folder_path)
    elif os.name == 'posix':
        os.system(f'open "{folder_path}"')

# ----------------------- API & NETWORK FUNCTIONS -----------------------
def search_google_images(title, search_suffix, api_key, cx, num_images=3):
    """
    Searches for images on Google using the Custom Search API.
    Increments the daily query counter and saves the new count.
    Returns a list of image URLs.
    """
    global daily_queries
    daily_queries += 1
    save_daily_queries(daily_queries)
    search_url = "https://www.googleapis.com/customsearch/v1"
    params = {
        'q': f"{title} {search_suffix}",
        'cx': cx,
        'key': api_key,
        'searchType': 'image',
        'num': num_images
    }
    response = requests.get(search_url, params=params)
    response.raise_for_status()
    results = response.json()
    return [item['link'] for item in results.get('items', [])]

def download_image(url, save_path):
    """
    Downloads an image from a URL and saves it to the given path.
    """
    response = requests.get(url)
    response.raise_for_status()
    with open(save_path, 'wb') as f:
        f.write(response.content)

# ----------------------- PARSING FUNCTIONS -----------------------
def parse_title_author(text):
    """
    Parses a string to extract a title and author.
    Supported delimiters: "--", "-", "/", "\\", "|"
    Returns a tuple (title, author) where author may be empty.
    """
    delimiters = ["--", "-", "/", "\\", "|"]
    pos = len(text)
    selected_delim = None
    for delim in delimiters:
        i = text.find(delim)
        if i != -1 and i < pos:
            pos = i
            selected_delim = delim
    if selected_delim:
        parts = text.split(selected_delim, 1)
        title = parts[0].strip()
        author = parts[1].strip()
        return title, author
    return text.strip(), ""

# ----------------------- DATABASE & IMAGE LOOKUP / SELECTION FUNCTIONS -----------------------
def check_database_for_title(category, title, author=""):
    """
    Searches the database for an image matching the given title (and optionally author).
    For Books: if an author is provided, attempt to return an exact match (by comparing normalized title and author).
    If no exact match is found (or no author provided), collect all matches based on normalized title.
    If multiple matches exist for Books, call a popup to select the correct version.
    For other categories, returns the matching image (or a selection popup if needed).
    """
    subfolder = CATEGORY_MAPPING.get(category)
    db_folder = os.path.join(DATABASE_FOLDER, subfolder)
    if not os.path.exists(db_folder):
        return None

    matches = []
    search_title = "".join(normalize_title(title).split()).lower()
    search_author = "".join(author.replace(".", "").split()).lower() if author else ""

    for file in os.listdir(db_folder):
        if file.lower().endswith(('.jpg', '.jpeg', '.png')):
            base = os.path.splitext(file)[0]
            if category == "Books":
                parts = base.split("_")
                file_title = parts[0].lower()
                if search_author and len(parts) >= 2:
                    file_author = parts[1].lower()
                    if file_title == search_title and file_author == search_author:
                        return os.path.join(db_folder, file)
                elif file_title == search_title:
                    matches.append(os.path.join(db_folder, file))
            else:
                if "".join(normalize_title(base).split()).lower() == search_title:
                    matches.append(os.path.join(db_folder, file))

    if not matches:
        return None
    if category == "Books" and len(matches) > 1:
        return select_book_version(matches, title)
    else:
        return matches[0] if len(matches) == 1 else select_database_image(matches, title)

def get_book_image_path(book_key):
    """
    Returns the full path for a book image in the database that matches the given book key.
    """
    books_folder = os.path.join(DATABASE_FOLDER, "Books")
    if not os.path.exists(books_folder):
        return None
    for file in os.listdir(books_folder):
        if file.lower().endswith(('.jpg', '.jpeg', '.png')):
            base = os.path.splitext(file)[0]
            if base.startswith(book_key):
                return os.path.join(books_folder, file)
    return None

def select_book_version(matches, title):
    """
    Creates a popup for the user to select one of multiple book versions found in the database.
    Also offers an option to add all versions.
    Returns the selected image path(s).
    """
    popup = tk.Toplevel()
    popup.title("Select Book Version")
    apply_background(popup)
    frame = tk.Frame(popup, bg="white")
    frame.pack(expand=True, fill="both", padx=10, pady=10)
    
    tk.Label(frame, text=f"Multiple versions found for '{title}'", 
             font=("Helvetica", int(14 * SCALE)), bg="white").pack(pady=10)
    tk.Label(frame, text="Select one to add to Gathered Images, or select 'Add All Versions'", 
             font=("Helvetica", int(12 * SCALE)), bg="white", wraplength=580, justify="center").pack(pady=5)
    
    var = tk.StringVar(value="")

    for idx, match in enumerate(matches, start=1):
        row_frame = tk.Frame(frame, bg="white")
        row_frame.pack(fill="x", pady=5)
        try:
            img = Image.open(match)
            img.thumbnail((100, 100))
            tk_img = ImageTk.PhotoImage(img)
            img_label = tk.Label(row_frame, image=tk_img, bg="white")
            img_label.image = tk_img
            img_label.pack(side="left", padx=5)
        except Exception as e:
            print(f"Error loading image {match}: {e}")
            tk.Label(row_frame, text="Image not available", bg="white").pack(side="left", padx=5)
        base = os.path.splitext(os.path.basename(match))[0]
        parts = base.split("_")
        if len(parts) >= 2:
            disp_title = parts[0]
            disp_author = parts[1]
            display_text = f"{disp_title} by {disp_author}"
        else:
            display_text = base
        tk.Radiobutton(row_frame, text=display_text, variable=var, value=match,
                       font=("Helvetica", int(12 * SCALE)), bg="white").pack(side="left", padx=10)
    
    tk.Radiobutton(frame, text="Add All Versions", variable=var, value="ALL",
                   font=("Helvetica", int(12 * SCALE)), bg="white").pack(pady=5)
    
    def confirm():
        if var.get() == "":
            messagebox.showinfo("Selection Required", "Please select an option.")
        else:
            popup.destroy()
    tk.Button(frame, text="Confirm", command=confirm,
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=int(15 * SCALE)).pack(pady=10)
    
    popup.wait_window()
    return matches if var.get() == "ALL" else var.get()

def select_database_image(matches, title):
    """
    Presents a popup for the user to select one image when multiple images are found in the database.
    Returns the selected image path.
    """
    result = None
    win = tk.Toplevel()
    win.title("Select Database Image")
    apply_background(win)
    container = tk.Frame(win, bg="white", bd=2)
    container.pack(expand=True, fill="both", padx=10, pady=10)
    tk.Label(container, text=f"Multiple images found for '{title}'.\nSelect one to copy:",
             font=("Helvetica", int(12 * SCALE)), bg="white", wraplength=380, justify="center").pack(pady=5)
    btn_frame = tk.Frame(container, bg="white")
    btn_frame.pack(pady=5)
    def on_select(path):
        nonlocal result
        result = path
        win.destroy()
    for image_path in matches:
        try:
            img = Image.open(image_path)
            img.thumbnail((int(100 * SCALE), int(100 * SCALE)))
            tk_img = ImageTk.PhotoImage(img)
            btn = tk.Button(btn_frame, image=tk_img, command=lambda p=image_path: on_select(p),
                            bg="#f7acb0", activebackground="lightpink")
            btn.image = tk_img
            btn.pack(side="left", padx=5, pady=5)
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
    win.wait_window()
    return result

def select_book_key(matches):
    """
    When multiple book metadata keys match a search, present a list for the user to choose one.
    Returns the selected book key.
    """
    popup = tk.Toplevel()
    popup.title("Select Book")
    apply_background(popup)
    tk.Label(popup, text="Multiple books found. Please select one:", font=("Helvetica", 14), bg="white").pack(pady=10)
    listbox = tk.Listbox(popup, font=("Helvetica", 12), width=40, height=10)
    for key in matches:
        listbox.insert(tk.END, key)
    listbox.pack(pady=5)
    selected_key = [None]
    def choose():
        try:
            index = listbox.curselection()[0]
            selected_key[0] = listbox.get(index)
            popup.destroy()
        except IndexError:
            messagebox.showinfo("Selection Required", "Please select a book.")
    tk.Button(popup, text="Select", command=choose, font=("Helvetica", 12), bg="#f7acb0", activebackground="lightpink", width=25).pack(pady=10)
    popup.wait_window()
    return selected_key[0]

# ----------------------- METADATA SEARCH & EDIT FUNCTIONS -----------------------
def search_book_metadata(author_search, selected_genres, page_from, page_to, copy_from, copy_to):
    """
    Searches the book metadata based on filters: author, genres, page count, and publication date.
    Returns a list of matching book keys.
    """
    metadata = load_book_metadata()
    results = []
    for key, data in metadata.items():
        if author_search and ("author" not in data or author_search.lower() not in data["author"].lower()):
            continue
        if selected_genres:
            book_genres = [g.lower() for g in data.get("genres", [])]
            if not all(genre.lower() in book_genres for genre in selected_genres):
                continue
        try:
            count = int(data.get("page_count", "0"))
        except:
            count = 0
        if page_from and count < int(page_from):
            continue
        if page_to and count > int(page_to):
            continue
        try:
            cdate = int(data.get("publication_date", "0"))
        except:
            cdate = 0
        if copy_from and cdate < int(copy_from):
            continue
        if copy_to and cdate > int(copy_to):
            continue
        results.append(key)
    return results

def display_search_results(matching_keys):
    """
    Displays search results for book metadata in a scrollable popup.
    Allows the user to select books to add to the Gathered Images folder.
    """
    if not matching_keys:
        messagebox.showinfo("Search Results", "No matching books found.")
        return
    popup = tk.Toplevel()
    popup.title("Search Results")
    apply_background(popup)
    container = tk.Frame(popup, bg="white", bd=2)
    container.pack(expand=True, fill="both", padx=10, pady=10)
    canvas = tk.Canvas(container, bg="white")
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")
    canvas.configure(yscrollcommand=scrollbar.set)
    results_frame = tk.Frame(canvas, bg="white")
    canvas.create_window((0,0), window=results_frame, anchor="nw")
    results_frame.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
    selected_books = {}
    def toggle_selection(book_key, btn):
        if book_key in selected_books:
            del selected_books[book_key]
            btn.config(relief="raised")
        else:
            selected_books[book_key] = btn
            btn.config(relief="sunken")
    max_cols = 5
    for idx, key in enumerate(matching_keys):
        row = idx // max_cols
        col = idx % max_cols
        img_path = get_book_image_path(key)
        if img_path:
            try:
                img = Image.open(img_path)
                img.thumbnail((int(100 * SCALE), int(100 * SCALE)))
                tk_img = ImageTk.PhotoImage(img)
                btn = tk.Button(results_frame, image=tk_img, bg="#f7acb0", activebackground="lightpink")
                btn.image = tk_img
                btn.config(command=lambda k=key, b=btn: toggle_selection(k, b))
                btn.grid(row=row, column=col, padx=5, pady=5)
            except Exception as e:
                print(f"Error loading image for '{key}': {e}")
        else:
            btn = tk.Button(results_frame, text=key, bg="#f7acb0", activebackground="lightpink")
            btn.config(command=lambda k=key, b=btn: toggle_selection(k, b))
            btn.grid(row=row, column=col, padx=5, pady=5)
    def add_selected():
        if not selected_books:
            messagebox.showinfo("No Selection", "No books selected.")
            return
        added_count = 0
        for book_key in selected_books.keys():
            img_path = get_book_image_path(book_key)
            if img_path:
                ext = os.path.splitext(img_path)[1]
                dest_filename = f"{book_key}{ext}"
                dest_path = os.path.join(GATHERED_IMAGES_FOLDER, dest_filename)
                try:
                    shutil.copy(img_path, dest_path)
                    added_count += 1
                except Exception as e:
                    print(f"Error copying image for '{book_key}': {e}")
        messagebox.showinfo("Success", f"Added {added_count} books to Gathered Images.")
        popup.destroy()
    add_button = tk.Button(popup, text="Add Selected Books to Gathered Images", command=add_selected,
                           font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=int(35 * SCALE))
    add_button.pack(pady=10)
    popup.wait_window()

def search_by_metadata():
    """
    Opens a popup that allows the user to search the book metadata using various criteria.
    """
    popup = tk.Toplevel()
    popup.title("Search Book Metadata")
    apply_background(popup, 1500, 1500)
    tk.Label(popup, text="Search Book Metadata", font=("Helvetica", int(14 * SCALE)), bg="white")\
        .pack(pady=10)
    
    # Author search input
    author_frame = tk.Frame(popup, bg="white")
    author_frame.pack(pady=5)
    tk.Label(author_frame, text="Author (contains):", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=0, column=0, sticky="e", padx=5, pady=5)
    author_entry = tk.Entry(author_frame, width=25)
    author_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    
    # Genre selection
    genre_frame = tk.Frame(popup, bg="white")
    genre_frame.pack(pady=5)
    tk.Label(genre_frame, text="Genres:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=0, column=0, sticky="ne", padx=5, pady=5)
    genres_frame = tk.Frame(genre_frame, bg="white")
    genres_frame.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    saved_genres = get_saved_genres()
    genre_vars = {}
    if saved_genres:
        total = len(saved_genres)
        rows = 7
        cols = math.ceil(total / rows)
        for i, genre in enumerate(saved_genres):
            r = i % rows
            c = i // rows
            var = tk.IntVar()
            chk = tk.Checkbutton(genres_frame, text=genre, variable=var,
                                 font=("Helvetica", int(12 * SCALE)), bg="white")
            chk.grid(row=r, column=c, padx=3, pady=3, sticky="w")
            genre_vars[genre] = var
    else:
        tk.Label(genres_frame, text="(No genres saved yet)", font=("Helvetica", int(12 * SCALE)), bg="white")\
            .pack()
    
    # Page count range inputs
    page_frame = tk.Frame(popup, bg="white")
    page_frame.pack(pady=5)
    tk.Label(page_frame, text="Page Count Range:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=0, column=0, sticky="e", padx=5, pady=5)
    page_from_entry = tk.Entry(page_frame, width=10)
    page_from_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    tk.Label(page_frame, text="to", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=0, column=2, padx=5, pady=5)
    page_to_entry = tk.Entry(page_frame, width=10)
    page_to_entry.grid(row=0, column=3, sticky="e", padx=5, pady=5)
    
    # Publication date range inputs
    publication_frame = tk.Frame(popup, bg="white")
    publication_frame.pack(pady=5)
    tk.Label(publication_frame, text="publication Year Range:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=0, column=0, sticky="e", padx=5, pady=5)
    copy_from_entry = tk.Entry(publication_frame, width=10)
    copy_from_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    tk.Label(publication_frame, text="to", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=0, column=2, padx=5, pady=5)
    copy_to_entry = tk.Entry(publication_frame, width=10)
    copy_to_entry.grid(row=0, column=3, sticky="e", padx=5, pady=5)
    
    def perform_search():
        author_search = author_entry.get().strip()
        selected = [genre for genre, var in genre_vars.items() if var.get() == 1]
        page_from = page_from_entry.get().strip()
        page_to = page_to_entry.get().strip()
        copy_from = copy_from_entry.get().strip()
        copy_to = copy_to_entry.get().strip()
        matching_keys = search_book_metadata(author_search, selected, page_from, page_to, copy_from, copy_to)
        popup.destroy()
        display_search_results(matching_keys)
        
    tk.Button(popup, text="Search", command=perform_search, 
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink")\
        .pack(pady=10)
    popup.wait_window()

def import_metadata_from_excel():
    """
    Imports book metadata from an Excel file.
    Expected columns: "title", "author", "publication date", "page count", and "genres".
    The metadata is stored in the JSON file with keys as composite keys from title and author.
    If any of the metadata fields (except title) are blank, they are stored as empty strings.
    """
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if not file_path:
        return
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Read header row and create a mapping from header name (lowercase) to column index.
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {str(cell).strip().lower(): idx for idx, cell in enumerate(header_row) if cell is not None}

    # Ensure that required columns exist.
    if "title" not in headers or "author" not in headers:
        messagebox.showerror("Import Error", "Excel file must contain 'title' and 'author' columns.")
        return

    meta = load_book_metadata()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Skip rows with an empty title.
        title_cell = row[headers["title"]]
        if title_cell is None or str(title_cell).strip() == "":
            continue
        title = str(title_cell).strip()
        author = str(row[headers["author"]]).strip() if row[headers["author"]] is not None else ""
        publication_date = (
            str(row[headers["publication date"]]).strip()
            if "publication date" in headers and row[headers["publication date"]] is not None
            else ""
        )
        page_count = (
            str(row[headers["page count"]]).strip()
            if "page count" in headers and row[headers["page count"]] is not None
            else ""
        )
        genres_str = (
            str(row[headers["genres"]]).strip()
            if "genres" in headers and row[headers["genres"]] is not None
            else ""
        )
        genres = [g.strip() for g in genres_str.split(",") if g.strip()]

        # Use the composite_key() function to generate the key (handles missing author appropriately)
        key = composite_key(title, author)
        meta[key] = {
            "title": title,
            "author": author,
            "publication_date": publication_date,
            "page_count": page_count,
            "genres": genres
        }
    save_book_metadata(meta)
    messagebox.showinfo("Import Successful", "Metadata successfully imported from Excel.")

SKIP_ALL_METADATA = False
def add_book_metadata_popup(source_path, book_title, default_author=""):
    """
    Opens a popup to add metadata for a book.
    A Title field is included so that the original (spaced) input is visible.
    The values entered by the user are retrieved via StringVar.
    When saving, the original spaced title and author are stored in the metadata block.
    If the user chooses "Skip All", this global flag is set and for all subsequent books
    the function returns a default metadata block (with empty publication date, page count, and genres).
    """
    global SKIP_ALL_METADATA
    # If skipping all metadata has already been chosen, immediately return default metadata.
    if SKIP_ALL_METADATA:
        input_title = book_title.strip()
        input_author = default_author.strip()
        if not input_author:
            messagebox.showerror("Missing Author", "Author is required for correct file naming. Please provide an author.")
            return None
        return {
            "title": input_title,
            "author": input_author,
            "publication_date": "",
            "page_count": "",
            "genres": []
        }
    
    popup = tk.Toplevel()
    popup.title("Add Metadata for Book")
    apply_background(popup)
    
    frame = tk.Frame(popup, bg="white")
    frame.pack(expand=True, fill="both", padx=10, pady=10)
    
    tk.Label(frame, text=f"Metadata for '{book_title}'", font=("Helvetica", int(16 * SCALE)), bg="white")\
        .grid(row=0, column=0, columnspan=2, pady=10)
    
    try:
        img = Image.open(source_path)
        img.thumbnail((int(200 * SCALE), int(200 * SCALE)))
        tk_img = ImageTk.PhotoImage(img)
        thumb_label = tk.Label(frame, image=tk_img, bg="white")
        thumb_label.image = tk_img
        thumb_label.grid(row=1, column=0, columnspan=2, pady=10)
    except Exception as e:
        print("Error loading image for metadata popup:", e)
    
    # Use StringVar so that changes are tracked.
    title_var = tk.StringVar(value=book_title)
    author_var = tk.StringVar(value=default_author)
    
    tk.Label(frame, text="Title:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=2, column=0, sticky="e", padx=5, pady=5)
    title_entry = tk.Entry(frame, width=30, font=("Helvetica", int(12 * SCALE)), textvariable=title_var)
    title_entry.grid(row=2, column=1, sticky="w", padx=5, pady=5)
    
    tk.Label(frame, text="Author:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=3, column=0, sticky="e", padx=5, pady=5)
    author_entry = tk.Entry(frame, width=30, font=("Helvetica", int(12 * SCALE)), textvariable=author_var)
    author_entry.grid(row=3, column=1, sticky="w", padx=5, pady=5)
    
    tk.Label(frame, text="publication date:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=4, column=0, sticky="e", padx=5, pady=5)
    publication_entry = tk.Entry(frame, width=30, font=("Helvetica", int(12 * SCALE)))
    publication_entry.grid(row=4, column=1, sticky="w", padx=5, pady=5)
    
    tk.Label(frame, text="Page Count:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=5, column=0, sticky="e", padx=5, pady=5)
    page_count_entry = tk.Entry(frame, width=30, font=("Helvetica", int(12 * SCALE)))
    page_count_entry.grid(row=5, column=1, sticky="w", padx=5, pady=5)
    
    tk.Label(frame, text="Existing Genres:", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=6, column=0, columnspan=2, pady=(10,5))
    saved_genres = get_saved_genres()
    genre_vars = {}
    if saved_genres:
        genres_frame = tk.Frame(frame, bg="white")
        genres_frame.grid(row=7, column=0, columnspan=2, pady=5)
        for i, genre in enumerate(saved_genres):
            var = tk.IntVar()
            chk = tk.Checkbutton(genres_frame, text=genre, variable=var,
                                 font=("Helvetica", int(12 * SCALE)), bg="white")
            chk.grid(row=i//4, column=i%4, padx=3, pady=3, sticky="w")
            genre_vars[genre] = var
    else:
        tk.Label(frame, text="(No existing genres)", font=("Helvetica", int(12 * SCALE)), bg="white")\
            .grid(row=7, column=0, columnspan=2, pady=5)
    
    tk.Label(frame, text="New Genres (comma-separated):", font=("Helvetica", int(12 * SCALE)), bg="white")\
        .grid(row=8, column=0, columnspan=2, pady=(10,5))
    new_genres_entry = tk.Entry(frame, width=35, font=("Helvetica", int(12 * SCALE)))
    new_genres_entry.grid(row=9, column=0, columnspan=2, pady=5)
    
    metadata_result = {}
    
    def save_metadata():
        input_title = title_var.get().strip()
        input_author = author_var.get().strip()
        metadata_result["title"] = input_title
        metadata_result["author"] = input_author
        metadata_result["publication_date"] = publication_entry.get().strip()
        metadata_result["page_count"] = page_count_entry.get().strip()
        selected = [genre for genre, var in genre_vars.items() if var.get() == 1]
        new = [g.strip() for g in new_genres_entry.get().split(",") if g.strip()]
        metadata_result["genres"] = selected + new
        popup.destroy()
    
    def skip_all_metadata():
        nonlocal metadata_result
        input_title = title_var.get().strip()
        input_author = author_var.get().strip()
        if not input_author:
            messagebox.showerror("Missing Author", "Author is required for correct file naming. Please provide an author before skipping.")
            return
        # Set the global flag so that all subsequent books are processed with default metadata.
        global SKIP_ALL_METADATA
        SKIP_ALL_METADATA = True
        metadata_result["title"] = input_title
        metadata_result["author"] = input_author
        metadata_result["publication_date"] = ""
        metadata_result["page_count"] = ""
        metadata_result["genres"] = []
        popup.destroy()
    
    btn_frame = tk.Frame(frame, bg="white")
    btn_frame.grid(row=10, column=0, columnspan=2, pady=20)
    
    tk.Button(btn_frame, text="Save Metadata", command=save_metadata, 
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=int(15 * SCALE))\
        .pack(side="left", padx=5)
    tk.Button(btn_frame, text="Skip All", command=skip_all_metadata,
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=int(15 * SCALE))\
        .pack(side="left", padx=5)
    
    popup.wait_window()
    return metadata_result if metadata_result else None

def open_edit_metadata_popup(book_key, current_metadata):
    """
    Opens a popup to edit metadata for a specific book.
    A Title field is included so the user can see and modify the normalized title.
    Upon saving, the metadata is stored using normalized values.
    """
    edit_popup = tk.Toplevel()
    edit_popup.title(f"Edit Metadata for '{book_key}'")
    apply_background(edit_popup, 1000, 1000)
    
    header_label = tk.Label(edit_popup, text=f"Current metadata for '{book_key}'",
                            font=("Helvetica", 18, "bold"), bg="white", fg="#f7acb0")
    header_label.grid(row=0, column=0, columnspan=2, pady=10)
    
    cover_frame = tk.Frame(edit_popup, bg="white")
    cover_frame.grid(row=1, column=0, columnspan=2, pady=10)
    img_path = get_book_image_path(book_key)
    if img_path:
        try:
            img = Image.open(img_path)
            img.thumbnail((150, 150))
            tk_img = ImageTk.PhotoImage(img)
            cover_label = tk.Label(cover_frame, image=tk_img, bg="white")
            cover_label.image = tk_img
            cover_label.pack()
        except Exception as e:
            print(f"Error loading image for '{book_key}': {e}")
    
    tk.Label(edit_popup, text="Title:", font=("Helvetica", 14), bg="white")\
        .grid(row=2, column=0, sticky="e", padx=10, pady=5)
    title_entry = tk.Entry(edit_popup, width=30, font=("Helvetica", 14))
    title_entry.grid(row=2, column=1, sticky="w", padx=10, pady=5)
    title_entry.insert(0, current_metadata.get("title", ""))
    
    tk.Label(edit_popup, text="Author:", font=("Helvetica", 14), bg="white")\
        .grid(row=3, column=0, sticky="e", padx=10, pady=5)
    author_entry = tk.Entry(edit_popup, width=30, font=("Helvetica", 14))
    author_entry.grid(row=3, column=1, sticky="w", padx=10, pady=5)
    author_entry.insert(0, current_metadata.get("author", ""))
    
    tk.Label(edit_popup, text="publication date:", font=("Helvetica", 14), bg="white")\
        .grid(row=4, column=0, sticky="e", padx=10, pady=5)
    publication_entry = tk.Entry(edit_popup, width=30, font=("Helvetica", 14))
    publication_entry.grid(row=4, column=1, sticky="w", padx=10, pady=5)
    publication_entry.insert(0, current_metadata.get("publication_date", ""))
    
    tk.Label(edit_popup, text="Page Count:", font=("Helvetica", 14), bg="white")\
        .grid(row=5, column=0, sticky="e", padx=10, pady=5)
    page_count_entry = tk.Entry(edit_popup, width=30, font=("Helvetica", 14))
    page_count_entry.grid(row=5, column=1, sticky="w", padx=10, pady=5)
    page_count_entry.insert(0, current_metadata.get("page_count", ""))
    
    genres_title = tk.Label(edit_popup, text="Genres", font=("Helvetica", 14), bg="white")
    genres_title.grid(row=6, column=0, columnspan=2, pady=(15,5))
    genres_frame = tk.Frame(edit_popup, bg="white")
    genres_frame.grid(row=7, column=0, columnspan=2, padx=10, pady=5)
    saved = set(get_saved_genres())
    current = set(current_metadata.get("genres", []))
    all_genres = sorted(list(saved.union(current)))
    genre_vars = {}
    cols = 3
    for i, genre in enumerate(all_genres):
        r = i // cols
        c = i % cols
        var = tk.IntVar(value=1 if genre in current else 0)
        chk = tk.Checkbutton(genres_frame, text=genre, variable=var, bg="white",
                             font=("Helvetica", 12))
        chk.grid(row=r, column=c, padx=5, pady=5, sticky="w")
        genre_vars[genre] = var
    
    new_genres_frame = tk.Frame(edit_popup, bg="white")
    new_genres_frame.grid(row=8, column=0, columnspan=2, padx=10, pady=5)
    new_genres_label = tk.Label(new_genres_frame, text="Add New Genres (comma-separated):",
                                font=("Helvetica", 14), bg="white")
    new_genres_label.pack(side="left", padx=5)
    new_genres_entry = tk.Entry(new_genres_frame, width=30, font=("Helvetica", 14))
    new_genres_entry.pack(side="left", padx=5)
    
    buttons_frame = tk.Frame(edit_popup, bg="white")
    buttons_frame.grid(row=9, column=0, columnspan=2, pady=15)
    
    def save_changes():
        normalized_title = "".join(normalize_title(title_entry.get().strip()).split())
        normalized_author = "".join(author_entry.get().strip().split())
        new_metadata = {
            "title": normalized_title,
            "author": normalized_author,
            "publication_date": publication_entry.get().strip(),
            "page_count": page_count_entry.get().strip(),
            "genres": [genre for genre, var in genre_vars.items() if var.get() == 1]
        }
        new_input = new_genres_entry.get().strip()
        if new_input:
            new_list = [g.strip() for g in new_input.split(",") if g.strip()]
            new_metadata["genres"].extend(new_list)
        meta = load_book_metadata()
        meta[book_key] = new_metadata
        save_book_metadata(meta)
        messagebox.showinfo("Success", f"Metadata for '{book_key}' updated.")
        edit_popup.destroy()
    
    tk.Button(buttons_frame, text="Save Changes", command=save_changes,
              font=("Helvetica", 14), bg="#f7acb0", activebackground="lightpink", width=15)\
        .pack(side="left", padx=10)
    tk.Button(buttons_frame, text="Cancel", command=edit_popup.destroy,
              font=("Helvetica", 14), bg="#f7acb0", activebackground="lightpink", width=15)\
        .pack(side="left", padx=10)
    edit_popup.wait_window()

# ----------------------- MAIN APPLICATION FUNCTIONS -----------------------
def send_images_to_database():
    """
    Processes images in the Gathered Images folder and moves them to the appropriate
    database folder based on their category. For Books, if a matching database image is found,
    it copies that image. Otherwise, it prompts for metadata.
    When prompting for metadata, the function retrieves the original spaced title and author
    (stored in ORIGINAL_BOOK_INPUTS) to prefill the metadata popup.
    The metadata is then stored using normalized title and author.
    """
    moved_any = False
    book_metadata = load_book_metadata()
    for root_dir, _, files in os.walk(GATHERED_IMAGES_FOLDER):
        for file in files:
            if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                file_path = os.path.join(root_dir, file)
                if "_db" in file:
                    os.remove(file_path)
                    continue
                for safe_cat, db_subfolder in CATEGORY_MAPPING.items():
                    if safe_cat == "Books":
                        if f"_{safe_cat}" not in file:
                            continue
                        # Expected filename format: "MyBook_AuthorName_[Books]..."
                        parts = file.split(f"_{safe_cat}")
                        original_key = parts[0]  # This is the normalized composite key.
                        dest_folder = os.path.join(DATABASE_FOLDER, db_subfolder)
                        os.makedirs(dest_folder, exist_ok=True)
                        extension = os.path.splitext(file)[1]
                        
                        db_image = check_database_for_title("Books", original_key)
                        if db_image:
                            dest_path = os.path.join(GATHERED_IMAGES_FOLDER, os.path.basename(db_image))
                            try:
                                shutil.copy(db_image, dest_path)
                                print(f"Copied {db_image} to {dest_path}")
                            except Exception as e:
                                print(e)
                        else:
                            if original_key in ORIGINAL_BOOK_INPUTS:
                                default_title, default_author = ORIGINAL_BOOK_INPUTS[original_key]
                            else:
                                if "_" in original_key:
                                    default_title, default_author = original_key.split("_", 1)
                                else:
                                    default_title = original_key
                                    default_author = ""
                            metadata = add_book_metadata_popup(file_path, default_title, default_author)
                            if metadata:
                                new_key = composite_key(metadata["title"], metadata["author"])
                                book_metadata[new_key] = metadata
                                save_book_metadata(book_metadata)
                                destination_path = os.path.join(dest_folder, new_key + extension)
                            else:
                                destination_path = None
                                break
                    else:
                        if f"_{safe_cat.replace(' ', '_')}" in file:
                            parts = file.split(f"_{safe_cat.replace(' ', '_')}")
                            title_part = parts[0]
                            extension = os.path.splitext(file)[1]
                            base_filename = "".join(normalize_title(title_part).split())
                            dest_folder = os.path.join(DATABASE_FOLDER, db_subfolder)
                            os.makedirs(dest_folder, exist_ok=True)
                            destination_path = os.path.join(dest_folder, base_filename + extension)
                        else:
                            continue
                    if destination_path:
                        try:
                            shutil.move(file_path, destination_path)
                            print(f"Moved {file} to {destination_path}")
                            moved_any = True
                        except Exception as e:
                            print(f"Error moving {file}: {e}")
                        break
    if moved_any:
        messagebox.showinfo("Transfer Complete", "Images have been sent to the database.")
    else:
        messagebox.showinfo("No Files Moved", "No matching images were found to transfer.")

def gather_media_covers_general(parent):
    """
    Opens the main window for gathering media covers.
    Provides options to enter titles for selected categories, downloads images,
    and allows selection of covers for books.
    """
    # Nested helper to auto-resize a text widget based on content.
    def auto_resize_text(event):
        widget = event.widget
        widget.update_idletasks()
        font = tkFont.Font(font=widget['font'])
        content = widget.get("1.0", "end-1c")
        avg_char_width = font.measure("a")
        widget_width_pixels = int(widget['width']) * avg_char_width
        lines = content.splitlines() if content else [""]
        total_lines = 0
        for line in lines:
            if not line:
                total_lines += 1
            else:
                line_width = font.measure(line)
                wrapped_lines = math.ceil(line_width / widget_width_pixels)
                total_lines += wrapped_lines
        widget.config(height=total_lines if total_lines > 0 else 1)

    window = tk.Toplevel(parent)
    window.title("Gather Media Covers")
    window.geometry("700x1000")
    WindowWidth = 700
    WindowHeight = 1000
    canvas = tk.Canvas(window, width=WindowWidth, height=WindowHeight, highlightthickness=0)
    canvas.place(x=0, y=0, relwidth=1, relheight=1)
    bg_image_path = os.path.join(UI_FOLDER, "flower background.png")
    try:
        bg_img = Image.open(bg_image_path)
        bg_img = bg_img.resize((WindowWidth, WindowHeight), Image.Resampling.LANCZOS)
        bg_photo = ImageTk.PhotoImage(bg_img)
        canvas.create_image(0, 0, image=bg_photo, anchor="nw")
        canvas.bg_photo = bg_photo
    except Exception as e:
        print("Error loading flower background image:", e)
    title_image_path = os.path.join(UI_FOLDER, "ImageCollector.png")
    try:
        title_img = Image.open(title_image_path)
        title_width = 500
        title_height = 100
        title_img = title_img.resize((title_width, title_height), Image.Resampling.LANCZOS)
        title_photo = ImageTk.PhotoImage(title_img)
        canvas.create_image(WindowWidth/2, 50, image=title_photo)
        canvas.title_photo = title_photo
    except Exception as e:
        print("Error loading title image:", e)

    # Category selection
    category_frame = tk.Frame(window, bg="white")
    category_frame.pack(anchor="n", padx=20, pady=(125,5))
    tk.Label(category_frame, text="Collect images for:", font=("Helvetica", int(12 * SCALE)), bg="white").pack(anchor="n")
    categories = [("Books", "Books"), ("Movies", "Movies"),
                  ("Video Games", "Video Games"), ("Albums", "Albums")]
    cat_vars = {}
    input_entries = {}
    input_container = tk.Frame(window, bg="white")
    input_container.pack(anchor="n", padx=20, pady=5)
    def update_category_input(cat):
        if cat_vars[cat].get():
            if cat not in input_entries:
                frame = tk.Frame(input_container, bg="white")
                frame.pack(anchor="w", pady=5)
                tk.Label(frame, text=f"{cat} titles:", font=("Helvetica", int(12 * SCALE)), bg="white").pack(side="left")
                text_widget = tk.Text(frame, width=40, height=1, wrap="word", font=("Helvetica", 12))
                text_widget.pack(side="left", padx=10)
                text_widget.bind("<KeyRelease>", auto_resize_text)
                input_entries[cat] = text_widget
            if not input_container.winfo_ismapped():
                input_container.place(x=20, y=240)
        else:
            if cat in input_entries:
                input_entries[cat].master.destroy()
                del input_entries[cat]
            if not input_entries:
                input_container.place_forget()
    for display, cat in categories:
        var = tk.BooleanVar(value=False)
        cat_vars[cat] = var
        cb = tk.Checkbutton(category_frame, text=display, variable=var, font=("Helvetica", int(12 * SCALE)), bg="white")
        cb.pack(side="left", padx=10)
        var.trace_add("write", lambda *args, cat=cat: update_category_input(cat))
    tk.Label(window, text="Friendly Reminder: Separate titles by commas and dont forget Authors!", font=("Helvetica", int(10 * SCALE)), bg="white").pack(pady=10)
    api_label = tk.Label(window, text=f"Current daily API queries: {daily_queries}", font=("Helvetica", int(10 * SCALE), "bold"), bg="white")
    api_label.pack(pady=5)
    def update_api_label():
        api_label.config(text=f"Current daily API queries: {daily_queries}")
    update_api_label()

    # Frame for displaying submission results
    result_frame = tk.Frame(window, bg="white")
    result_frame.pack(pady=10)
    # Frames for the Submit and Send-to-Database buttons
    submit_frame = tk.Frame(window, bg="white")
    submit_frame.pack(pady=5)
    S2Database_frame = tk.Frame(window, bg="white")
    S2Database_frame.pack(pady=5)

    # Helper to allow selection of book covers from a temporary folder.
    def select_book_covers(temp_folder, columns=6):
        selected_files = []
        popup = tk.Toplevel(window)
        popup.title("Select Book Covers")
        frame = tk.Frame(popup, bg="white")
        frame.pack(fill="both", expand=True, padx=10, pady=10)
        tk.Label(frame, text="Select covers to keep (click to toggle selection):",
                 font=("Helvetica", int(12 * SCALE)), bg="white").pack(pady=5)
        grid_frame = tk.Frame(frame, bg="white")
        grid_frame.pack()
        image_files = [os.path.join(temp_folder, f) for f in os.listdir(temp_folder)
                       if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
        selection_state = {}
        def toggle_selection(file, btn):
            current = selection_state.get(file, False)
            new_state = not current
            selection_state[file] = new_state
            if new_state:
                btn.config(relief="sunken", borderwidth=3)
            else:
                btn.config(relief="raised", borderwidth=1)
        thumb_size = (int(100 * SCALE), int(100 * SCALE))
        row = 0
        col = 0
        for file in image_files:
            try:
                img = Image.open(file)
                img.thumbnail(thumb_size)
                tk_img = ImageTk.PhotoImage(img)
                btn = tk.Button(grid_frame, image=tk_img, relief="raised", borderwidth=1,
                                bg="#f7acb0", activebackground="lightpink")
                btn.image = tk_img
                btn.config(command=lambda f=file, b=btn: toggle_selection(f, b))
                btn.grid(row=row, column=col, padx=5, pady=5)
                selection_state[file] = False
                col += 1
                if col >= columns:
                    col = 0
                    row += 1
            except Exception as e:
                print(f"Error loading image {file}: {e}")
        def confirm_selection():
            for file, selected in selection_state.items():
                if selected:
                    selected_files.append(file)
            popup.destroy()
        tk.Button(frame, text="Confirm Selection", command=confirm_selection,
                  font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink").pack(pady=10)
        popup.wait_window()
        return selected_files

    def submit():
        for widget in result_frame.winfo_children():
            widget.destroy()
        data = {}
        # Process input for each selected category.
        for cat, var in cat_vars.items():
            if var.get() and cat in input_entries:
                text = input_entries[cat].get("1.0", "end-1c")
                items = [i.strip() for i in text.split(',') if i.strip()]
                parsed_items = []
                for raw_item in items:
                    parsed_title, parsed_author = parse_title_author(raw_item)
                    if cat == "Books":
                        key = composite_key(parsed_title, parsed_author)
                        parsed_items.append((parsed_title, parsed_author, key))
                        ORIGINAL_BOOK_INPUTS[key] = (parsed_title, parsed_author)
                    else:
                        parsed_items.append((normalize_title(parsed_title), parsed_author))
                if parsed_items:
                    data[cat] = parsed_items
        if not data:
            messagebox.showinfo("Input Error", "Please select at least one category and enter titles (for Books, author is optional).")
            return
        os.makedirs(GATHERED_IMAGES_FOLDER, exist_ok=True)
        search_suffixes = {
            'Books': "book cover",
            'Movies': "movie poster",
            'Video Games': "video game cover",
            'Albums': "album cover"
        }
        processed_book_titles = []
        temp_books_folder = os.path.join("temp_books")
        os.makedirs(temp_books_folder, exist_ok=True)
        for category, items in data.items():
            for item in items:
                if category == "Books":
                    title, author, book_key = item
                    processed_book_titles.append((title, author))
                    dest_folder = os.path.join(DATABASE_FOLDER, CATEGORY_MAPPING.get(category))
                    os.makedirs(dest_folder, exist_ok=True)
                    db_image = check_database_for_title("Books", title, author)
                    if db_image:
                        base = os.path.splitext(os.path.basename(db_image))[0] + "_db"
                        ext = os.path.splitext(db_image)[1]
                        dest_filename = base + ext
                        dest_path = os.path.join(GATHERED_IMAGES_FOLDER, dest_filename)
                        try:
                            shutil.copy(db_image, dest_path)
                            print(f"Copied {db_image} to {dest_path}")
                        except Exception as e:
                            print(e)
                    else:
                        try:
                            query = f"{title} by {author}" if author else title
                            image_links = search_google_images(query, search_suffixes.get(category, ""), API_KEY, CX, num_images=3)
                            update_api_label()
                            if not image_links:
                                print(f"No images found for {title} in {category}")
                                continue
                            for idx, image_url in enumerate(image_links):
                                file_name = f"{book_key}_{category}_{idx+1}.jpg"
                                save_path = os.path.join(temp_books_folder, file_name)
                                download_image(image_url, save_path)
                                print(f"Downloaded {file_name} from {image_url} to temporary folder")
                        except Exception as e:
                            print(f"Error downloading images for {title} in {category}: {e}")
                else:
                    title, author = item
                    safe_title = "".join(normalize_title(title).split())
                    safe_category = category.replace(" ", "_")
                    dest_folder = os.path.join(DATABASE_FOLDER, CATEGORY_MAPPING.get(category))
                    os.makedirs(dest_folder, exist_ok=True)
                    try:
                        query = f"{title} by {author}" if author else title
                        db_image = check_database_for_title(category, title)
                        if db_image:
                            ext = os.path.splitext(db_image)[1]
                            file_name = f"{safe_title}_{safe_category}_db{ext}"
                            dest_path = os.path.join(GATHERED_IMAGES_FOLDER, file_name)
                            shutil.copy(db_image, dest_path)
                            print(f"Copied {db_image} to {dest_path}")
                        else:
                            image_links = search_google_images(query, search_suffixes.get(category, ""), API_KEY, CX, num_images=3)
                            update_api_label()
                            if not image_links:
                                print(f"No images found for {title} in {category}")
                                continue
                            for idx, image_url in enumerate(image_links):
                                file_name = f"{safe_title}_{safe_category}_{idx+1}.jpg"
                                save_path = os.path.join(GATHERED_IMAGES_FOLDER, file_name)
                                download_image(image_url, save_path)
                                print(f"Downloaded {file_name} from {image_url}")
                    except Exception as e:
                        print(f"Error processing {title} in {category}: {e}")
        if os.path.exists(temp_books_folder) and os.listdir(temp_books_folder):
            selected_files = select_book_covers(temp_books_folder, columns=6)
            added_count = 0
            for file_path in selected_files:
                file_name = os.path.basename(file_path)
                dest_path = os.path.join(GATHERED_IMAGES_FOLDER, file_name)
                try:
                    shutil.copy(file_path, dest_path)
                    print(f"Copied {file_path} to {dest_path}")
                    added_count += 1
                except Exception as e:
                    print(f"Error copying selected cover {file_path}: {e}")
            messagebox.showinfo("Selection Complete", f"Added {added_count} book covers to Gathered Images.")
            try:
                shutil.rmtree(temp_books_folder)
            except Exception as e:
                print(f"Error removing temporary folder: {e}")
        if processed_book_titles:
            list_str = "\n".join(f"{i+1}. {title}" + (f" by {author}" if author else "") for i, (title, author) in enumerate(processed_book_titles))
            summary_label = tk.Label(result_frame, text="Books Searched:\n" + list_str, font=("Helvetica", int(12 * SCALE)), bg="white", justify="left")
            summary_label.pack()
        else:
            summary_label = tk.Label(result_frame, text="No book titles were processed.", font=("Helvetica", int(12 * SCALE)), bg="white")
            summary_label.pack()

    tk.Button(submit_frame, text="Submit", command=submit, font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=20).pack()
    tk.Button(S2Database_frame, text="Send to Database", command=send_images_to_database, font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=20).pack()
    window.wait_window()

def edit_metadata_options():
    """
    Opens a popup offering options to edit metadata:
      - Edit metadata for a specific title,
      - Input missing metadata,
      - Import metadata from an Excel file.
    """
    popup = tk.Toplevel()
    popup.title("Edit Metadata Options")
    popup.geometry("400x300")
    apply_background(popup, 500, 500)
    header = tk.Label(popup, text="Edit Metadata Options",
                      font=("Helvetica", int(16 * SCALE)), bg="white")
    header.pack(pady=15)
    def edit_specific():
        popup.destroy()
        edit_book_metadata()
    def input_missing():
        popup.destroy()
        key = select_missing_metadata_book()
        if key:
            meta = load_book_metadata()
            open_edit_metadata_popup(key, meta[key])
    def import_excel():
        popup.destroy()
        import_metadata_from_excel()
    tk.Button(popup, text="Edit Metadata for Specific Title", 
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=25,
              command=edit_specific).pack(pady=5)
    tk.Button(popup, text="Input Missing Metadata", 
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=25,
              command=input_missing).pack(pady=5)
    tk.Button(popup, text="Input Excel File", 
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=25,
              command=import_excel).pack(pady=5)
    popup.wait_window()

def edit_book_metadata():
    """
    Opens a popup to allow the user to search for a book by title and then edit its metadata.
    """
    popup = tk.Toplevel()
    popup.title("Edit Book Metadata")
    apply_background(popup)
    frame = tk.Frame(popup, bg="white")
    frame.pack(expand=True, fill="both", padx=10, pady=10)
    tk.Label(frame, text="Enter the title of the book to edit:", 
             font=("Helvetica", int(12 * SCALE)), bg="white").pack(pady=5)
    title_entry = tk.Entry(frame, width=40)
    title_entry.pack(pady=5)
    def search_title():
        input_title = title_entry.get().strip()
        if not input_title:
            messagebox.showinfo("Error", "Please enter a title.")
            return
        meta = load_book_metadata()
        norm_input = "".join(normalize_title(input_title).split())
        matches = [key for key in meta.keys() if key.startswith(norm_input)]
        if not matches:
            messagebox.showinfo("Not Found", f"No book with title '{input_title}' found in metadata.")
            popup.destroy()
        elif len(matches) == 1:
            found_key = matches[0]
            popup.destroy()
            open_edit_metadata_popup(found_key, meta[found_key])
        else:
            found_key = select_book_key(matches)
            if found_key:
                popup.destroy()
                open_edit_metadata_popup(found_key, meta[found_key])
    tk.Button(frame, text="Search", command=search_title, 
              font=("Helvetica", int(12 * SCALE)), bg="#f7acb0", activebackground="lightpink", width=int(15 * SCALE))\
        .pack(pady=10)
    popup.wait_window()

def get_books_with_missing_metadata():
    meta = load_book_metadata()
    missing = []
    for key, data in meta.items():
        if (not data.get("author") or not data.get("publication_date") or 
            not data.get("page_count") or not data.get("genres")):
            missing.append(key)
    return missing

def export_missing_metadata_to_excel():
    """
    Creates an Excel file with column headers for each metadata category
    and writes one row per book from the missing metadata list.
    Each row contains: Composite Key, Title, Author, publication date, Page Count, Genres.
    The file is saved using a save-as dialog.
    """
    missing_books = get_books_with_missing_metadata()
    if not missing_books:
        messagebox.showinfo("No Missing Metadata", "No books with missing metadata found.")
        return
    meta = load_book_metadata()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missing Metadata"
    headers = ["Title", "Author", "publication date", "Page Count", "Genres"]
    ws.append(headers)
    for key in missing_books:
        data = meta.get(key, {})
        genres = ", ".join(data.get("genres", []))
        row = [data.get("title", ""),
               data.get("author", ""),
               data.get("publication_date", ""),
               data.get("page_count", ""),
               genres]
        ws.append(row)
    save_path = filedialog.asksaveasfilename(
        title="Save Missing Metadata",
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")]
    )
    if save_path:
        try:
            wb.save(save_path)
            messagebox.showinfo("Export Successful", f"Missing metadata exported to:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save file: {e}")

def select_missing_metadata_book():
    """
    Opens a window listing all books with missing metadata.
    Next to the "Edit Selected" button, an "Export to Excel" button is provided.
    The "Edit Selected" button returns the selected book key for editing,
    while the "Export to Excel" button saves the missing metadata list to an Excel file.
    """
    popup = tk.Toplevel()
    popup.title("Select Book with Missing Metadata")
    apply_background(popup)
    frame = tk.Frame(popup, bg="white")
    frame.pack(expand=True, fill="both", padx=10, pady=10)
    tk.Label(frame, text="Books with Missing Metadata", font=("Helvetica", 14), bg="white").pack(pady=10)
    missing_books = get_books_with_missing_metadata()
    if not missing_books:
        messagebox.showinfo("No Missing Metadata", "No books with missing metadata were found.")
        popup.destroy()
        return None
    listbox = tk.Listbox(frame, font=("Helvetica", 12), width=40, height=10)
    for book in missing_books:
        listbox.insert(tk.END, book)
    listbox.pack(pady=5)
    selected_key = [None]
    def select_book():
        try:
            index = listbox.curselection()[0]
            selected_key[0] = listbox.get(index)
            popup.destroy()
        except IndexError:
            messagebox.showinfo("Selection Required", "Please select a book.")
    button_frame = tk.Frame(frame, bg="white")
    button_frame.pack(pady=10)
    tk.Button(button_frame, text="Edit Selected", command=select_book,
              font=("Helvetica", 12), bg="#f7acb0", activebackground="lightpink", width=20)\
        .pack(side="left", padx=5)
    tk.Button(button_frame, text="Export to Excel", command=export_missing_metadata_to_excel,
              font=("Helvetica", 12), bg="#f7acb0", activebackground="lightpink", width=20)\
        .pack(side="left", padx=5)
    popup.wait_window()
    return selected_key[0]

def main_menu():
    """
    Main menu for the Media Cover Manager application.
    Provides options to gather media covers, search/edit metadata, and open relevant folders.
    """
    root = tk.Tk()
    root.title("Media Cover Manager")
    WindowWidth = 550
    WindowHeight = 500
    root.geometry(f"{WindowWidth}x{WindowHeight}")
    canvas = tk.Canvas(root, width=WindowWidth, height=WindowHeight, highlightthickness=0)
    canvas.place(x=0, y=0, relwidth=1, relheight=1)
    bg_image_path = os.path.join(UI_FOLDER, "flower background.png")
    try:
        bg_img = Image.open(bg_image_path)
        bg_img = bg_img.resize((WindowWidth, WindowHeight), Image.Resampling.LANCZOS)
        bg_photo = ImageTk.PhotoImage(bg_img)
        canvas.create_image(0, 0, image=bg_photo, anchor="nw")
        canvas.bg_photo = bg_photo
    except Exception as e:
        print("Error loading flower background image:", e)
    title_image_path = os.path.join(UI_FOLDER, "MegsMediaCollector.png")
    try:
        title_img = Image.open(title_image_path)
        title_width = 500
        title_height = 100
        title_img = title_img.resize((title_width, title_height), Image.Resampling.LANCZOS)
        title_photo = ImageTk.PhotoImage(title_img)
        canvas.create_image(WindowWidth/2, 50, image=title_photo)
        canvas.title_photo = title_photo
    except Exception as e:
        print("Error loading title image:", e)
    buttons = [
        ("Gather Media Covers", lambda: gather_media_covers_general(root)),
        ("Search by Metadata", search_by_metadata),
        ("Edit Metadata", lambda: edit_metadata_options()),
        ("Open Gathered Images Folder", lambda: open_file_explorer(GATHERED_IMAGES_FOLDER)),
        ("Open Database Folder", lambda: open_file_explorer(DATABASE_FOLDER))
    ]
    button_width = 350
    button_height = 50
    start_y = 125
    vertical_spacing = button_height + 10
    x_coord = (WindowWidth - button_width) / 2
    for idx, (text, command) in enumerate(buttons):
        y_coord = start_y + idx * vertical_spacing
        btn = tk.Button(root, text=text, command=command,
                        font=("Helvetica", int(12 * SCALE)),
                        bg="#f7acb0", activebackground="lightpink")
        btn.place(x=x_coord, y=y_coord, width=button_width, height=button_height)
    root.mainloop()

if __name__ == "__main__":
    main_menu()
